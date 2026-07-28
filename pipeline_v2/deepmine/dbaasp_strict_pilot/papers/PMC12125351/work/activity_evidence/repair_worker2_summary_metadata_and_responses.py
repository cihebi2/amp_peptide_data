#!/usr/bin/env python3
"""Repair worker-2 summary metadata and append nonterminal repair responses.

This script is paper-local for PMC12125351. It reads existing row-level
worker-2 activity/toxicity evidence, validates locator/source-cell bindings in
aggregate, recomputes summary metadata from accepted records, writes the worker
and final mirrors, and appends one fresh owner response for each assigned
worker-2 rework ticket. Stdout is limited to aggregate status.
"""

from __future__ import annotations

import json
import math
import re
import unicodedata
from collections import Counter
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


PAPER_ID = "PMC12125351"
WORKER = "worker-2"
SOURCE_WORKBOOK_NAME = "42003_2025_8282_MOESM2_ESM.xlsx"

ROOT = Path(__file__).resolve().parents[4]
PAPER_ROOT = ROOT / "papers" / PAPER_ID
PACKET_ROOT = ROOT / "packets" / PAPER_ID
WORK_DIR = PAPER_ROOT / "work" / "activity_evidence"

PAPER_WORK_OUT = WORK_DIR / "activity_records.json"
PACKET_WORKER2_OUT = PACKET_ROOT / "analysis" / "activity_toxicity_evidence.worker2.json"
PAPER_FINAL_OUT = PAPER_ROOT / "final" / "activity_toxicity_evidence.json"
PACKET_FINAL_OUT = PACKET_ROOT / "final" / "activity_toxicity_evidence.json"
VALIDATION_OUT = WORK_DIR / "worker2_summary_metadata_repair_validation.json"
APPEND_SUMMARY_OUT = WORK_DIR / "worker2_summary_metadata_rework_response_append_summary.json"
RESPONSE_OUT = PACKET_ROOT / "rework" / "rework_responses.jsonl"
WORKBOOK = PACKET_ROOT / "raw" / "supplementary_original" / SOURCE_WORKBOOK_NAME
LOCATOR_INDEX = PACKET_ROOT / "locators" / "locator_index.json"

NOW = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

SHEET_ORDER = [
    "Supplementary Data 3",
    "Supplementary Data 4",
    "Supplementary Data 10",
    "Supplementary Data 11",
    "Supplementary Data 12",
]

ASSIGNED_TICKETS = [
    "rwk-PMC12125351-campaign-r01-BF-PMC12125351-W2-ACTIVITY-TOXICITY-UNDEREXTRACTED",
    "rwk-PMC12125351-campaign-r01-BF-PMC12125351-W2-SD10-STRAIN-CONFLICT-METADATA",
    "rwk-PMC12125351-campaign-r02-BF-PMC12125351-W2-ACTIVITY-HARD-FINDING-NOT-RECONCILED",
    "rwk-PMC12125351-campaign-r02-BF-PMC12125351-W2-ACTIVITY-TOXICITY-SOURCE-FIELD-CONFLICTS",
    "rwk-PMC12125351-campaign-r03-BF-PMC12125351-W2-ACTIVITY-SUMMARY-METADATA-PLACEHOLDER",
    "rwk-PMC12125351-campaign-r03-BF-PMC12125351-W2-ACTIVITY-TOXICITY-FIELD-CONFLICTS",
]

EXPECTED_SHEET_COUNTS = {
    ("activity_records", "Supplementary Data 3", "MIC"): 76,
    ("activity_records", "Supplementary Data 4", "MIC"): 36,
    ("activity_records", "Supplementary Data 10", "MIC"): 18,
    ("toxicity_records", "Supplementary Data 10", "CC50/HC50"): 18,
    ("toxicity_records", "Supplementary Data 11", "percent hemolysis"): 54,
    ("toxicity_records", "Supplementary Data 12", "cell viability"): 54,
}

P17_P20_IDS = {
    "PMC12125351-SD4-R006-C05-MIC": {"value_cell": "E6", "parallel_cell": "F6"},
    "PMC12125351-SD4-R007-C05-MIC": {"value_cell": "E7", "parallel_cell": "F7"},
}

ALLOWED_NORMALIZATION = {"direct", "converted", "not_convertible", "ambiguous"}


def load_json(path: Path) -> Any:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2, sort_keys=False)
        handle.write("\n")


def rel(path: Path) -> str:
    return str(path.relative_to(ROOT))


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        return format(value, ".15g")
    return " ".join(unicodedata.normalize("NFKC", str(value)).split())


def try_float(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = norm_text(value)
    if not re.fullmatch(r"[<>]?\s*-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", text):
        return None
    return float(text.replace("<", "").replace(">", "").strip())


def values_equal(left: Any, right: Any) -> bool:
    left_float = try_float(left)
    right_float = try_float(right)
    if left_float is not None and right_float is not None:
        return abs(left_float - right_float) <= 1e-9
    return norm_text(left) == norm_text(right)


def parse_locator(locator: str | None) -> dict[str, Any] | None:
    if not locator:
        return None
    sheet_match = re.search(r"sheet=([^:]+)", locator)
    row_match = re.search(r":row=(\d+)", locator)
    cell_match = re.search(r":cell=([A-Z]+)(\d+)$", locator)
    if not (sheet_match and row_match and cell_match):
        return None
    return {
        "sheet": sheet_match.group(1),
        "row": int(row_match.group(1)),
        "cell": f"{cell_match.group(1)}{cell_match.group(2)}",
        "column": cell_match.group(1),
        "column_index": column_index_from_string(cell_match.group(1)),
    }


def sheet_name(locator: str | None) -> str | None:
    match = re.search(r"sheet=([^:]+)", locator or "")
    return match.group(1) if match else None


def source_row(locator: str | None) -> int | None:
    match = re.search(r":row=(\d+)", locator or "")
    return int(match.group(1)) if match else None


def load_locator_set() -> set[str]:
    data = load_json(LOCATOR_INDEX)
    locators: set[str] = set()
    for item in data.get("locators", []):
        if isinstance(item, dict) and isinstance(item.get("locator"), str):
            locators.add(item["locator"])
        elif isinstance(item, str):
            locators.add(item)
    return locators


def locator_strings_from_field(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value] if value.startswith(("supp:", "xml:", "pdf:", "database:")) else []
    if isinstance(value, list):
        out: list[str] = []
        for item in value:
            out.extend(locator_strings_from_field(item))
        return out
    if isinstance(value, dict):
        out = []
        for item in value.values():
            out.extend(locator_strings_from_field(item))
        return out
    return []


def collect_record_locators(record: dict[str, Any]) -> list[str]:
    locators: list[str] = []
    for key, value in record.items():
        if "locator" in key or key in {"source_conflicts", "preserved_source_conflict", "worker2_ticket_repair"}:
            locators.extend(locator_strings_from_field(value))
    return locators


def endpoint_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    return dict(Counter(str(row.get("endpoint") or "not_reported") for row in rows))


def summarize_sheet_locators(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for sheet in SHEET_ORDER:
        sheet_rows = [row for row in rows if sheet_name(row.get("source_locator")) == sheet]
        if not sheet_rows:
            continue
        row_numbers = [row_num for row_num in (source_row(row.get("source_locator")) for row in sheet_rows) if row_num]
        summary[sheet] = {
            "record_count": len(sheet_rows),
            "cell_locator_count": len({row.get("source_locator") for row in sheet_rows if row.get("source_locator")}),
            "source_locator_prefix": f"supp:{SOURCE_WORKBOOK_NAME}:sheet={sheet}",
            "source_row_min": min(row_numbers) if row_numbers else None,
            "source_row_max": max(row_numbers) if row_numbers else None,
            "endpoint_counts": endpoint_counts(sheet_rows),
        }
    return summary


def recompute_summary(data: dict[str, Any]) -> dict[str, Any]:
    activity_records = data.get("activity_records", [])
    toxicity_records = data.get("toxicity_records", [])
    all_rows = activity_records + toxicity_records
    all_endpoint_counts = endpoint_counts(all_rows)
    activity_locators = summarize_sheet_locators(activity_records)
    toxicity_locators = summarize_sheet_locators(toxicity_records)

    source_role_counts = {
        "supplementary_data_3_activity": activity_locators.get("Supplementary Data 3", {}).get("record_count", 0),
        "supplementary_data_4_activity": activity_locators.get("Supplementary Data 4", {}).get("record_count", 0),
        "supplementary_data_10_activity_log10_mic": activity_locators.get("Supplementary Data 10", {}).get("record_count", 0),
        "supplementary_data_10_toxicity_log10_cc50_hc50": toxicity_locators.get("Supplementary Data 10", {}).get("record_count", 0),
        "supplementary_data_11_hemolysis": toxicity_locators.get("Supplementary Data 11", {}).get("record_count", 0),
        "supplementary_data_12_cell_viability": toxicity_locators.get("Supplementary Data 12", {}).get("record_count", 0),
    }
    accepted_source_tables = []
    for sheet in SHEET_ORDER:
        roles = []
        if sheet in activity_locators:
            roles.append("activity")
        if sheet in toxicity_locators:
            roles.append("toxicity")
        if roles:
            accepted_source_tables.append(
                {
                    "sheet": sheet,
                    "evidence_roles": roles,
                    "record_count": activity_locators.get(sheet, {}).get("record_count", 0)
                    + toxicity_locators.get(sheet, {}).get("record_count", 0),
                    "source_locator_prefix": f"supp:{SOURCE_WORKBOOK_NAME}:sheet={sheet}",
                }
            )

    summary = deepcopy(data.get("summary_counts") or {})
    for legacy_key in (
        "activity_tables_accepted",
        "accepted_activity_locators",
        "toxicity_tables_accepted",
        "accepted_toxicity_locators",
    ):
        summary.pop(legacy_key, None)
    summary.update(
        {
            "activity_records": len(activity_records),
            "toxicity_records": len(toxicity_records),
            "activity_source_sheets_accepted": len(activity_locators),
            "activity_tables_excluded": int(summary.get("activity_tables_excluded") or 0),
            "toxicity_source_sheets_accepted": len(toxicity_locators),
            "source_tables_checked": len(accepted_source_tables),
            "accepted_activity_source_sheets": activity_locators,
            "accepted_toxicity_source_sheets": toxicity_locators,
            "accepted_source_tables": accepted_source_tables,
            **source_role_counts,
            "data10_cc50_hc50_records": sum(
                1
                for row in toxicity_records
                if sheet_name(row.get("source_locator")) == "Supplementary Data 10"
                and row.get("endpoint") in {"CC50", "HC50"}
            ),
            "data10_log10_mic_records": source_role_counts["supplementary_data_10_activity_log10_mic"],
            "data10_selectivity_records": sum(1 for row in all_rows if row.get("endpoint") == "selectivity index"),
            "activity_exclusion_count": len(data.get("excluded_activity_source_cells_or_rows") or []),
        }
    )
    return summary


def source_role_counts_from_summary(summary: dict[str, Any]) -> dict[str, int]:
    keys = [
        "supplementary_data_3_activity",
        "supplementary_data_4_activity",
        "supplementary_data_10_activity_log10_mic",
        "supplementary_data_10_toxicity_log10_cc50_hc50",
        "supplementary_data_11_hemolysis",
        "supplementary_data_12_cell_viability",
    ]
    return {key: int(summary.get(key) or 0) for key in keys}


def update_artifact(data: dict[str, Any], validation_rel: str) -> dict[str, Any]:
    updated = deepcopy(data)
    summary = recompute_summary(updated)
    updated["summary_counts"] = summary
    updated["updated_at"] = NOW
    qa_summary = updated.setdefault("qa_summary", {})
    qa_summary["endpoint_counts"] = endpoint_counts(updated.get("activity_records", []) + updated.get("toxicity_records", []))
    qa_summary["source_role_counts"] = source_role_counts_from_summary(summary)
    qa_summary["normalization_status_counts"] = dict(
        Counter(row.get("normalization_status") for row in updated.get("activity_records", []) + updated.get("toxicity_records", []))
    )
    qa_summary["summary_metadata_recomputed"] = True
    qa_summary["summary_metadata_recomputed_at"] = NOW

    validation_artifacts = list(updated.get("validation_artifacts") or [])
    if validation_rel not in validation_artifacts:
        validation_artifacts.append(validation_rel)
    updated["validation_artifacts"] = validation_artifacts
    updated["worker2_ticket_repair_summary"] = {
        "repair_status": "repair_ready_for_adjudication",
        "repair_status_is_terminal": False,
        "worker": WORKER,
        "repaired_at": NOW,
        "assigned_ticket_count": len(ASSIGNED_TICKETS),
        "assigned_tickets": ASSIGNED_TICKETS,
        "summary_metadata_recomputed": True,
        "source_tables_checked": summary["source_tables_checked"],
        "activity_source_sheets_accepted": summary["activity_source_sheets_accepted"],
        "toxicity_source_sheets_accepted": summary["toxicity_source_sheets_accepted"],
        "validation_artifact": validation_rel,
        "terminal_adjudication": "deferred_to_worker_6",
    }
    return updated


def validate_rows(data: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    locators = load_locator_set()
    rows_by_group = {
        "activity_records": data.get("activity_records", []),
        "toxicity_records": data.get("toxicity_records", []),
    }
    issues: list[dict[str, Any]] = []
    scalar_checked = 0
    series_checked = 0
    locator_checked = 0
    unresolved_locators: list[dict[str, Any]] = []
    direct_conversion_conflicts = 0
    normalized_required_missing = 0
    concentration_conflicts = 0

    for group, rows in rows_by_group.items():
        for row in rows:
            record_id = row.get("record_id")
            status = row.get("normalization_status")
            if status not in ALLOWED_NORMALIZATION:
                issues.append({"code": "invalid_normalization_status", "record_id": record_id})
            if status in {"direct", "converted"} and (
                "normalized_value" not in row or "normalized_unit" not in row or row.get("normalized_unit") in {None, ""}
            ):
                normalized_required_missing += 1
                issues.append({"code": "normalized_fields_missing", "record_id": record_id})
            if status == "direct":
                if not values_equal(row.get("raw_value"), row.get("normalized_value")) or norm_text(row.get("raw_unit")) != norm_text(
                    row.get("normalized_unit")
                ):
                    direct_conversion_conflicts += 1
                    issues.append({"code": "direct_normalization_changes_value_or_unit", "record_id": record_id})
            if row.get("endpoint") in {"MIC", "MBC", "MFC", "IC50", "EC50", "HC50", "CC50"} and not row.get("raw_unit"):
                issues.append({"code": "unit_missing_for_quantitative_endpoint", "record_id": record_id})
            assay_conditions = row.get("assay_conditions") if isinstance(row.get("assay_conditions"), dict) else {}
            top_conc = row.get("concentration")
            nested_conc = assay_conditions.get("peptide_concentration") or assay_conditions.get("sample_concentration")
            if top_conc is not None and nested_conc is not None and not values_equal(top_conc, nested_conc):
                concentration_conflicts += 1
                issues.append({"code": "concentration_field_conflict", "record_id": record_id})

            for locator in sorted(set(collect_record_locators(row))):
                locator_checked += 1
                if locator not in locators:
                    unresolved_locators.append({"record_id": record_id, "locator": locator})

            parsed = parse_locator(row.get("source_locator"))
            raw_value = row.get("raw_value")
            if parsed and parsed["sheet"] in wb.sheetnames:
                ws = wb[parsed["sheet"]]
                if isinstance(raw_value, list):
                    same_row_value_locs = []
                    for locator in [row.get("source_locator")] + list(row.get("supporting_source_locators") or []):
                        item = parse_locator(locator)
                        if item and item["sheet"] == parsed["sheet"] and item["row"] == parsed["row"] and item["column"] in {"C", "D", "E"}:
                            same_row_value_locs.append(item)
                    same_row_value_locs.sort(key=lambda item: item["column_index"])
                    observed = [ws[item["cell"]].value for item in same_row_value_locs]
                    if len(observed) != len(raw_value) or any(not values_equal(left, right) for left, right in zip(raw_value, observed)):
                        issues.append({"code": "source_series_value_mismatch", "record_id": record_id, "source_locator": row.get("source_locator")})
                    series_checked += 1
                else:
                    observed = ws[parsed["cell"]].value
                    if not values_equal(raw_value, observed):
                        issues.append({"code": "source_cell_value_mismatch", "record_id": record_id, "source_locator": row.get("source_locator")})
                    scalar_checked += 1

    activity = rows_by_group["activity_records"]
    toxicity = rows_by_group["toxicity_records"]
    summary = recompute_summary(data)
    sheet_count_checks = {}
    for (group, sheet, endpoint), expected in EXPECTED_SHEET_COUNTS.items():
        rows = rows_by_group[group]
        if endpoint == "CC50/HC50":
            actual = sum(1 for row in rows if sheet_name(row.get("source_locator")) == sheet and row.get("endpoint") in {"CC50", "HC50"})
        else:
            actual = sum(1 for row in rows if sheet_name(row.get("source_locator")) == sheet and row.get("endpoint") == endpoint)
        sheet_count_checks[f"{group}:{sheet}:{endpoint}"] = {"expected": expected, "actual": actual, "passed": actual == expected}
        if actual != expected:
            issues.append({"code": "expected_sheet_count_mismatch", "group": group, "sheet": sheet, "endpoint": endpoint})

    ticket_checks = {
        "sd3_ecoli_k88_strain_not_reported_count": sum(
            1
            for row in activity
            if sheet_name(row.get("source_locator")) == "Supplementary Data 3"
            and row.get("target_species") == "Escherichia coli"
            and norm_text(row.get("target_strain_or_isolate")).casefold() == "not reported"
        ),
        "data10_toxicity_bad_log_unit_count": sum(
            1
            for row in toxicity
            if sheet_name(row.get("source_locator")) == "Supplementary Data 10"
            and str(row.get("raw_endpoint_label") or "").startswith("log10")
            and ("uM" in norm_text(row.get("raw_unit")) or "log2" in norm_text(row.get("raw_unit")))
        ),
        "data10_selectivity_toxicity_count": sum(
            1 for row in toxicity if sheet_name(row.get("source_locator")) == "Supplementary Data 10" and row.get("endpoint") == "selectivity index"
        ),
        "data10_to_12_homo_sapiens_toxicity_count": sum(
            1
            for row in toxicity
            if sheet_name(row.get("source_locator")) in {"Supplementary Data 10", "Supplementary Data 11", "Supplementary Data 12"}
            and row.get("target_species") == "Homo sapiens"
        ),
        "hc50_percent_hemolysis_bad_incubation_count": sum(
            1
            for row in toxicity
            if row.get("endpoint") in {"HC50", "percent hemolysis"}
            and (row.get("assay_conditions") or {}).get("incubation_time") != "1 h"
        ),
        "summary_source_tables_checked": summary.get("source_tables_checked"),
        "summary_activity_source_sheets_accepted": summary.get("activity_source_sheets_accepted"),
        "summary_accepted_activity_source_sheets_nonempty": bool(summary.get("accepted_activity_source_sheets")),
    }

    for key, value in ticket_checks.items():
        if key.endswith("_count") and value != 0:
            issues.append({"code": key, "count": value})
    if summary.get("source_tables_checked") != len({sheet_name(row.get("source_locator")) for row in activity + toxicity if sheet_name(row.get("source_locator"))}):
        issues.append({"code": "summary_source_tables_checked_mismatch"})
    if not summary.get("accepted_activity_source_sheets"):
        issues.append({"code": "summary_accepted_activity_source_sheets_empty"})

    p17_p20_checks = {}
    for record_id, cells in P17_P20_IDS.items():
        row = next((item for item in activity if item.get("record_id") == record_id), None)
        passed = False
        if row:
            conflict = row.get("preserved_source_conflict")
            parallels = row.get("source_reported_parallel_values") or []
            value_locator = f"supp:{SOURCE_WORKBOOK_NAME}:sheet=Supplementary Data 4:row={cells['value_cell'][1:]}:cell={cells['value_cell']}"
            parallel_locator = f"supp:{SOURCE_WORKBOOK_NAME}:sheet=Supplementary Data 4:row={cells['parallel_cell'][1:]}:cell={cells['parallel_cell']}"
            raw_match = row.get("source_locator") == value_locator
            parallel_match = any(item.get("source_locator") == parallel_locator for item in parallels if isinstance(item, dict))
            passed = bool(conflict and raw_match and parallel_match)
        p17_p20_checks[record_id] = {"passed": passed}
        if not passed:
            issues.append({"code": "p17_p20_parallel_conflict_missing", "record_id": record_id})

    blocker_codes = {
        item.get("code")
        for item in data.get("unresolved_blockers", [])
        if isinstance(item, dict)
    }
    stale_blocker_absent = "p17_p20_paeruginosa_um_and_xml_p24_conflict_not_preserved" not in blocker_codes
    if not stale_blocker_absent:
        issues.append({"code": "stale_p17_p20_unresolved_blocker_present"})

    sd10_column_e_checks = []
    sd10_column_e_failures = 0
    sd10_column_e_rows = [
        row
        for row in activity
        if sheet_name(row.get("source_locator")) == "Supplementary Data 10" and re.search(r":cell=E\d+$", row.get("source_locator") or "")
    ]
    for row in sd10_column_e_rows:
        conflicts = row.get("source_conflicts") or []
        label_locator_ok = True
        conflict_mentions_source = False
        conflict_mentions_value_basis = False
        no_column_locator = True
        for conflict in conflicts:
            if not isinstance(conflict, dict):
                continue
            label_locator = conflict.get("source_label_locator")
            if not label_locator or ":column=E" in label_locator or label_locator not in locators:
                label_locator_ok = False
            conflict_text = json.dumps(conflict, ensure_ascii=True)
            conflict_mentions_source = conflict_mentions_source or "ATCC 25923" in conflict_text
            conflict_mentions_value_basis = conflict_mentions_value_basis or "ATCC 29213" in conflict_text
            no_column_locator = no_column_locator and ":column=E" not in conflict_text
        row_passed = (
            "ATCC 25923" in str(row.get("raw_endpoint_label") or "")
            and label_locator_ok
            and no_column_locator
            and conflict_mentions_source
            and (row.get("target_strain_or_isolate") != "ATCC 29213" or conflict_mentions_value_basis)
        )
        if not row_passed:
            sd10_column_e_failures += 1
            issues.append({"code": "sd10_column_e_conflict_metadata_incomplete", "record_id": row.get("record_id")})
        sd10_column_e_checks.append({"record_id": row.get("record_id"), "passed": row_passed})

    return {
        "paper_id": PAPER_ID,
        "validated_at": NOW,
        "artifact_checked": rel(PACKET_WORKER2_OUT),
        "activity_count": len(activity),
        "toxicity_count": len(toxicity),
        "summary_counts_after_repair": {
            key: summary.get(key)
            for key in [
                "activity_records",
                "toxicity_records",
                "source_tables_checked",
                "activity_source_sheets_accepted",
                "toxicity_source_sheets_accepted",
                "data10_log10_mic_records",
                "data10_cc50_hc50_records",
                "data10_selectivity_records",
            ]
        },
        "expected_sheet_count_checks": sheet_count_checks,
        "workbook_value_checks": {
            "scalar_cells_checked": scalar_checked,
            "series_rows_checked": series_checked,
            "mismatch_count": sum(1 for issue in issues if issue["code"] in {"source_cell_value_mismatch", "source_series_value_mismatch"}),
        },
        "locator_resolution": {
            "locator_values_checked": locator_checked,
            "unresolved_count": len(unresolved_locators),
            "unresolved": unresolved_locators[:25],
        },
        "normalization_checks": {
            "allowed_status_values": sorted(ALLOWED_NORMALIZATION),
            "status_counts": dict(Counter(row.get("normalization_status") for row in activity + toxicity)),
            "direct_conversion_conflicts": direct_conversion_conflicts,
            "normalized_required_missing": normalized_required_missing,
            "concentration_conflicts": concentration_conflicts,
        },
        "ticket_contract_checks": {
            **ticket_checks,
            "p17_p20_checks": p17_p20_checks,
            "p17_p20_stale_blocker_absent": stale_blocker_absent,
            "sd10_column_e_rows_checked": len(sd10_column_e_rows),
            "sd10_column_e_failures": sd10_column_e_failures,
            "sd10_column_e_checks": sd10_column_e_checks,
        },
        "issue_count": len(issues) + len(unresolved_locators),
        "issues": issues[:50],
        "validation_status": "passed" if not issues and not unresolved_locators else "failed",
    }


def build_response(ticket_id: str, validation: dict[str, Any], artifacts_written: list[str]) -> dict[str, Any]:
    ticket_contract_checks = validation["ticket_contract_checks"]
    response = {
        "ticket_id": ticket_id,
        "response_status": "repair_ready_for_adjudication",
        "response_by": WORKER,
        "analysis_can_resume": True,
        "responded_at": NOW,
        "paper_id": PAPER_ID,
        "evidence": {
            "activity_records": validation["activity_count"],
            "toxicity_records": validation["toxicity_count"],
            "source_tables_checked": validation["summary_counts_after_repair"]["source_tables_checked"],
            "validation_status": validation["validation_status"],
            "issue_count": validation["issue_count"],
        },
        "evidence_paths": [rel(VALIDATION_OUT), rel(WORKBOOK), rel(LOCATOR_INDEX)],
        "repaired_artifacts": [
            rel(PAPER_WORK_OUT),
            rel(PACKET_WORKER2_OUT),
            rel(PAPER_FINAL_OUT),
            rel(PACKET_FINAL_OUT),
        ],
        "artifacts_written": artifacts_written,
        "added_files": [rel(Path(__file__)), rel(VALIDATION_OUT), rel(APPEND_SUMMARY_OUT)],
        "validation_artifacts": [rel(VALIDATION_OUT)],
        "reason": "Fresh worker-2 source-cell, locator, conflict-field, and summary-metadata validation passed; response is nonterminal and ready for worker-6 adjudication.",
        "notes": {
            "owner_response_is_terminal": False,
            "worker6_must_adjudicate": True,
            "expected_sheet_count_checks_passed": all(item["passed"] for item in validation["expected_sheet_count_checks"].values()),
            "locator_unresolved_count": validation["locator_resolution"]["unresolved_count"],
            "sd10_column_e_rows_checked": ticket_contract_checks["sd10_column_e_rows_checked"],
            "sd10_column_e_failures": ticket_contract_checks["sd10_column_e_failures"],
            "p17_p20_stale_blocker_absent": ticket_contract_checks["p17_p20_stale_blocker_absent"],
        },
    }
    return response


def main() -> int:
    data = load_json(PACKET_WORKER2_OUT)
    validation = validate_rows(data)
    validation_rel = rel(VALIDATION_OUT)

    artifacts_written: list[str] = []
    if validation["validation_status"] == "passed":
        for path in [PAPER_WORK_OUT, PACKET_WORKER2_OUT, PAPER_FINAL_OUT, PACKET_FINAL_OUT]:
            current = load_json(path)
            updated = update_artifact(current, validation_rel)
            write_json(path, updated)
            artifacts_written.append(rel(path))

        validation = validate_rows(load_json(PACKET_WORKER2_OUT))
        validation["artifacts_written"] = artifacts_written
        validation["mirror_consistency"] = {
            "paper_work_vs_packet_worker2": PAPER_WORK_OUT.read_bytes() == PACKET_WORKER2_OUT.read_bytes(),
            "paper_final_vs_packet_final": PAPER_FINAL_OUT.read_bytes() == PACKET_FINAL_OUT.read_bytes(),
        }
    else:
        validation["artifacts_written"] = []
        validation["mirror_consistency"] = {
            "paper_work_vs_packet_worker2": False,
            "paper_final_vs_packet_final": False,
        }

    write_json(VALIDATION_OUT, validation)

    responses_appended = []
    if validation["validation_status"] == "passed":
        RESPONSE_OUT.parent.mkdir(parents=True, exist_ok=True)
        with RESPONSE_OUT.open("a", encoding="utf-8") as handle:
            for ticket_id in ASSIGNED_TICKETS:
                response = build_response(ticket_id, validation, artifacts_written + [rel(VALIDATION_OUT), rel(APPEND_SUMMARY_OUT)])
                handle.write(json.dumps(response, ensure_ascii=False, sort_keys=False) + "\n")
                responses_appended.append(ticket_id)

    append_summary = {
        "paper_id": PAPER_ID,
        "worker": WORKER,
        "appended_at": NOW,
        "response_status": "repair_ready_for_adjudication" if validation["validation_status"] == "passed" else "validation_failed_no_response_appended",
        "responses_appended": responses_appended,
        "response_count": len(responses_appended),
        "validation_artifact": validation_rel,
        "artifacts_written": artifacts_written + [rel(VALIDATION_OUT), rel(APPEND_SUMMARY_OUT)],
    }
    write_json(APPEND_SUMMARY_OUT, append_summary)

    print(
        json.dumps(
            {
                "status": append_summary["response_status"],
                "activity_records": validation["activity_count"],
                "toxicity_records": validation["toxicity_count"],
                "source_tables_checked": validation["summary_counts_after_repair"]["source_tables_checked"],
                "issue_count": validation["issue_count"],
                "responses_appended": len(responses_appended),
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0 if validation["validation_status"] == "passed" else 1


if __name__ == "__main__":
    raise SystemExit(main())
