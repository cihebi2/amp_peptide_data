#!/usr/bin/env python3
"""Repair worker-2 activity/toxicity field conflicts for PMC12125351."""

from __future__ import annotations

import json
import math
import re
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PAPER_ID = "PMC12125351"
WORKER_ID = "worker-2"
TICKET_ID = "rwk-PMC12125351-campaign-r03-BF-PMC12125351-W2-ACTIVITY-TOXICITY-FIELD-CONFLICTS"

SCRIPT_PATH = Path(__file__).resolve()
REPO_ROOT = SCRIPT_PATH.parents[7]
BASE = REPO_ROOT / "pipeline_v2/deepmine/dbaasp_strict_pilot"
PAPER_ROOT = BASE / "papers" / PAPER_ID
PACKET_ROOT = BASE / "packets" / PAPER_ID
WORK_DIR = PAPER_ROOT / "work/activity_evidence"

WORK_ACTIVITY = WORK_DIR / "activity_records.json"
ANALYSIS_WORKER2 = PACKET_ROOT / "analysis/activity_toxicity_evidence.worker2.json"
PAPER_FINAL = PAPER_ROOT / "final/activity_toxicity_evidence.json"
PACKET_FINAL = PACKET_ROOT / "final/activity_toxicity_evidence.json"
REWORK_RESPONSES = PACKET_ROOT / "rework/rework_responses.jsonl"
XLSX_PATH = PAPER_ROOT / "source/supplementary/42003_2025_8282_MOESM2_ESM.xlsx"

ARTIFACT_PATHS = [WORK_ACTIVITY, ANALYSIS_WORKER2, PAPER_FINAL, PACKET_FINAL]

XLSX_LOC_RE = re.compile(
    r"^supp:42003_2025_8282_MOESM2_ESM\.xlsx:sheet=(?P<sheet>.+):row=(?P<row>\d+):cell=(?P<cell>[A-Z]+\d+)$"
)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def ensure_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def add_unique(values: list[Any], item: Any) -> list[Any]:
    if item not in values:
        values.append(item)
    return values


def supp_locator(sheet: str, row: int, cell: str | None = None) -> str:
    loc = f"supp:42003_2025_8282_MOESM2_ESM.xlsx:sheet={sheet}:row={row}"
    if cell:
        loc += f":cell={cell}"
    return loc


def row_number_from_sd10_e_locator(locator: str) -> int | None:
    match = XLSX_LOC_RE.match(locator)
    if not match:
        return None
    if match.group("sheet") != "Supplementary Data 10":
        return None
    row = int(match.group("row"))
    if match.group("cell") != f"E{row}":
        return None
    if 3 <= row <= 11:
        return row
    return None


def is_staph_data10_conflict_row(row: dict[str, Any]) -> bool:
    locator = str(row.get("source_locator", ""))
    return (
        row_number_from_sd10_e_locator(locator) is not None
        and row.get("endpoint") == "MIC"
        and str(row.get("target_species", "")).lower() == "staphylococcus aureus"
    )


def log10_match_locators(wb: Any, log_value: Any, sheet: str, column: str) -> list[str]:
    if not isinstance(log_value, (int, float)):
        return []
    ws = wb[sheet]
    matches = []
    for row_num in range(1, ws.max_row + 1):
        cell = f"{column}{row_num}"
        if log10_equal(log_value, ws[cell].value):
            matches.append(supp_locator(sheet, row_num, cell))
    return matches


def repair_object(obj: dict[str, Any], repaired_at: str, wb: Any) -> tuple[dict[str, Any], dict[str, int]]:
    obj = deepcopy(obj)
    counts = {
        "hemolysis_endpoint_rows_repaired": 0,
        "supp_data10_staph_conflict_rows_repaired": 0,
    }

    for row in obj.get("toxicity_records", []):
        if row.get("endpoint") not in {"HC50", "percent hemolysis"}:
            continue
        assay_conditions = row.setdefault("assay_conditions", {})
        assay_conditions["incubation_time"] = "1 h"
        assay_conditions["incubation_time_basis_locator"] = "xml:p:86"
        assay_conditions["method_basis_locator"] = "xml:p:86"
        method_locators = ensure_list(assay_conditions.get("method_locators"))
        assay_conditions["method_locators"] = add_unique(method_locators, "xml:p:86")
        supporting = ensure_list(row.get("supporting_source_locators"))
        row["supporting_source_locators"] = add_unique(supporting, "xml:p:86")
        row["worker2_ticket_repair"] = {
            "ticket_id": TICKET_ID,
            "repaired_at": repaired_at,
            "fields_repaired": [
                "assay_conditions.incubation_time",
                "assay_conditions.incubation_time_basis_locator",
                "assay_conditions.method_locators",
            ],
            "source_basis_locators": ["xml:p:86"],
            "repair_status": "repair_ready_for_adjudication",
        }
        counts["hemolysis_endpoint_rows_repaired"] += 1

    for row in obj.get("activity_records", []):
        if not is_staph_data10_conflict_row(row):
            continue
        sd10_row = row_number_from_sd10_e_locator(str(row.get("source_locator", "")))
        assert sd10_row is not None
        data3_row = sd10_row + 5
        data4_row = sd10_row + 1
        leader_hint_data3_locator = supp_locator("Supplementary Data 3", data3_row, f"N{data3_row}")
        leader_hint_data4_locator = supp_locator("Supplementary Data 4", data4_row, f"G{data4_row}")
        data10_value = cell_value(wb, "Supplementary Data 10", f"E{sd10_row}")
        data3_match_locators = log10_match_locators(wb, data10_value, "Supplementary Data 3", "N")
        data4_match_locators = log10_match_locators(wb, data10_value, "Supplementary Data 4", "G")

        previous_strain = row.get("target_strain_or_isolate")
        row["target_strain_or_isolate"] = "ATCC 29213"
        row["target_strain_or_isolate_source"] = "value_provenance_with_preserved_conflict"
        row["source_reported_target_strain_or_isolate"] = previous_strain
        row["value_provenance_target_strain_or_isolate"] = "ATCC 29213"
        row["target_strain_conflict_status"] = "source_conflict_preserved"
        row["source_conflicts"] = [
            {
                "conflict_type": "target_strain_label_value_provenance_divergence",
                "status": "source_conflict_preserved",
                "field": "target_strain_or_isolate",
                "source_label_value": previous_strain,
                "source_label_locator": "supp:42003_2025_8282_MOESM2_ESM.xlsx:sheet=Supplementary Data 10:column=E",
                "assigned_value": "ATCC 29213",
                "assigned_value_basis": "value_provenance",
                "value_provenance_locators": ["xml:p:25", "xml:caption:4", *data3_match_locators],
                "value_provenance_match_status": (
                    "single_data3_column_n_match"
                    if len(data3_match_locators) == 1
                    else "multiple_data3_column_n_candidate_matches"
                    if data3_match_locators
                    else "no_data3_column_n_numeric_match"
                ),
                "leader_hint_locators_checked": [leader_hint_data3_locator, leader_hint_data4_locator],
                "contrasting_label_candidate_locators": data4_match_locators,
                "contrasting_label_match_status": (
                    "numeric_overlap_with_data4_column_g_present"
                    if data4_match_locators
                    else "no_numeric_match_in_data4_column_g"
                ),
                "caution": "Supplementary Data 10 column E label and Data 3/Fig. 4 value provenance diverge; conflict preserved for worker-6 adjudication.",
            }
        ]
        supporting = ensure_list(row.get("supporting_source_locators"))
        for locator in ["xml:p:25", "xml:caption:4", *data3_match_locators, *data4_match_locators]:
            supporting = add_unique(supporting, locator)
        row["supporting_source_locators"] = supporting
        row["worker2_ticket_repair"] = {
            "ticket_id": TICKET_ID,
            "repaired_at": repaired_at,
            "fields_repaired": [
                "target_strain_or_isolate",
                "target_strain_or_isolate_source",
                "source_reported_target_strain_or_isolate",
                "value_provenance_target_strain_or_isolate",
                "source_conflicts",
            ],
            "source_basis_locators": ["xml:p:25", "xml:caption:4", *data3_match_locators, *data4_match_locators],
            "repair_status": "repair_ready_for_adjudication",
        }
        counts["supp_data10_staph_conflict_rows_repaired"] += 1

    obj["reviewed_at"] = repaired_at
    obj["source_review_status"] = "source_reviewed_worker_repair_ready_for_adjudication"
    obj["publication_grade_claim"] = False
    obj["publication_grade_rationale"] = (
        "Worker-2 repair is source-reviewed for the assigned activity/toxicity field ticket; "
        "terminal publication-grade closure remains worker-6 adjudication."
    )
    obj["worker2_rework_response_status"] = "repair_ready_for_adjudication"
    obj["worker2_ticket_repair_summary"] = {
        "ticket_id": TICKET_ID,
        "repaired_at": repaired_at,
        **counts,
        "analysis_can_resume": True,
    }
    validation_artifacts = ensure_list(obj.get("validation_artifacts"))
    for rel_path in [
        "papers/PMC12125351/work/activity_evidence/worker2_field_conflict_repair_validation.json",
        "papers/PMC12125351/work/activity_evidence/worker2_sd10_value_provenance_reconciliation.json",
        "papers/PMC12125351/work/activity_evidence/worker2_ticket_xml_section_locator_checks.json",
    ]:
        validation_artifacts = add_unique(validation_artifacts, rel_path)
    obj["validation_artifacts"] = validation_artifacts
    return obj, counts


def cell_value(wb: Any, sheet: str, cell: str) -> Any:
    return wb[sheet][cell].value


def values_equal(row_value: Any, source_value: Any) -> bool:
    if isinstance(row_value, (int, float)) and isinstance(source_value, (int, float)):
        return row_value == source_value
    return str(row_value) == str(source_value)


def row_raw_value_matches_workbook(row: dict[str, Any], wb: Any, sheet: str, row_num: int, cell: str) -> bool:
    raw_value = row.get("raw_value")
    source_value = cell_value(wb, sheet, cell)
    if not isinstance(raw_value, list):
        return values_equal(raw_value, source_value)

    source_cells = (row.get("statistics") or {}).get("source_value_cells") or []
    if len(raw_value) != len(source_cells):
        return False
    for index, cell_info in enumerate(source_cells):
        col = cell_info.get("cell")
        if not col:
            return False
        source_cell = f"{col}{row_num}"
        workbook_value = cell_value(wb, sheet, source_cell)
        if not values_equal(raw_value[index], workbook_value):
            return False
        if not values_equal(cell_info.get("value"), workbook_value):
            return False
    return values_equal(raw_value[0], source_value)


def log10_equal(log_value: Any, linear_value: Any) -> bool:
    if not isinstance(log_value, (int, float)) or not isinstance(linear_value, (int, float)):
        return False
    return math.isclose(float(log_value), math.log10(float(linear_value)), rel_tol=0.0, abs_tol=1e-12)


def validate_obj(obj: dict[str, Any], wb: Any) -> dict[str, Any]:
    activity = obj.get("activity_records", [])
    toxicity = obj.get("toxicity_records", [])
    rows = activity + toxicity
    norm_allowed = {"direct", "converted", "not_convertible", "ambiguous"}
    xlsx_rows = []
    xlsx_mismatch_ids = []
    invalid_norm_ids = []
    direct_conversion_conflict_ids = []
    concentration_conflict_ids = []
    hemo_bad_ids = []
    sd10_conflict_bad_ids = []
    sd10_reconciliation = []

    for row in rows:
        row_id = row.get("record_id")
        status = row.get("normalization_status")
        if status not in norm_allowed:
            invalid_norm_ids.append(row_id)
        if status in {"direct", "converted"} and (
            "normalized_value" not in row or "normalized_unit" not in row
        ):
            invalid_norm_ids.append(row_id)
        if status == "direct":
            if row.get("normalized_value") != row.get("raw_value") or row.get("normalized_unit") != row.get("raw_unit"):
                direct_conversion_conflict_ids.append(row_id)

        assay_conditions = row.get("assay_conditions") or {}
        if "concentration" in row or "concentration_unit" in row:
            if (
                row.get("concentration") != assay_conditions.get("peptide_concentration")
                or row.get("concentration_unit") != assay_conditions.get("peptide_concentration_unit")
            ):
                concentration_conflict_ids.append(row_id)

        if row.get("endpoint") in {"HC50", "percent hemolysis"}:
            if assay_conditions.get("incubation_time") != "1 h" or "xml:p:86" not in ensure_list(assay_conditions.get("method_locators")):
                hemo_bad_ids.append(row_id)

        locator = str(row.get("source_locator", ""))
        match = XLSX_LOC_RE.match(locator)
        if match:
            sheet = match.group("sheet")
            row_num = int(match.group("row"))
            cell = match.group("cell")
            xlsx_rows.append(row_id)
            if not row_raw_value_matches_workbook(row, wb, sheet, row_num, cell):
                xlsx_mismatch_ids.append(row_id)

        if is_staph_data10_conflict_row(row):
            sd10_row = row_number_from_sd10_e_locator(locator)
            assert sd10_row is not None
            data10_value = cell_value(wb, "Supplementary Data 10", f"E{sd10_row}")
            data3_match_locators = log10_match_locators(wb, data10_value, "Supplementary Data 3", "N")
            data4_match_locators = log10_match_locators(wb, data10_value, "Supplementary Data 4", "G")
            has_conflict = (
                row.get("target_strain_or_isolate") == "ATCC 29213"
                and row.get("target_strain_conflict_status") == "source_conflict_preserved"
                and row.get("source_conflicts")
            )
            if not has_conflict:
                sd10_conflict_bad_ids.append(row_id)
            sd10_reconciliation.append(
                {
                    "record_id": row_id,
                    "data10_locator": supp_locator("Supplementary Data 10", sd10_row, f"E{sd10_row}"),
                    "data3_log10_match_locators": data3_match_locators,
                    "data4_log10_match_locators": data4_match_locators,
                    "data3_log10_match_count": len(data3_match_locators),
                    "data4_log10_match_count": len(data4_match_locators),
                    "has_data3_column_n_value_provenance_match": bool(data3_match_locators),
                    "data4_column_g_checked_for_reconciliation": True,
                    "conflict_status": row.get("target_strain_conflict_status"),
                    "assigned_target_strain_or_isolate": row.get("target_strain_or_isolate"),
                    "source_reported_target_strain_or_isolate": row.get("source_reported_target_strain_or_isolate"),
                }
            )

    return {
        "activity_record_count": len(activity),
        "toxicity_record_count": len(toxicity),
        "xlsx_source_cell_rows_checked": len(xlsx_rows),
        "xlsx_source_cell_mismatch_count": len(xlsx_mismatch_ids),
        "xlsx_source_cell_mismatch_record_ids": xlsx_mismatch_ids,
        "invalid_normalization_count": len(set(invalid_norm_ids)),
        "invalid_normalization_record_ids": sorted(set(invalid_norm_ids)),
        "direct_conversion_conflict_count": len(direct_conversion_conflict_ids),
        "direct_conversion_conflict_record_ids": direct_conversion_conflict_ids,
        "concentration_conflict_count": len(concentration_conflict_ids),
        "concentration_conflict_record_ids": concentration_conflict_ids,
        "hemolysis_endpoint_rows_checked": sum(1 for row in rows if row.get("endpoint") in {"HC50", "percent hemolysis"}),
        "hemolysis_bad_record_ids": hemo_bad_ids,
        "supp_data10_staph_rows_checked": len(sd10_reconciliation),
        "supp_data10_staph_conflict_bad_record_ids": sd10_conflict_bad_ids,
        "sd10_value_provenance_reconciliation": sd10_reconciliation,
    }


def append_rework_response(repaired_at: str, validation_path: Path, reconciliation_path: Path) -> None:
    response = {
        "ticket_id": TICKET_ID,
        "response_status": "repair_ready_for_adjudication",
        "response_by": WORKER_ID,
        "analysis_can_resume": True,
        "responded_at": repaired_at,
        "paper_id": PAPER_ID,
        "reason": (
            "Repaired worker-2 activity/toxicity field conflicts for HC50/percent hemolysis "
            "incubation metadata and Supplementary Data 10 Staphylococcus strain provenance."
        ),
        "evidence": [
            {
                "check": "HC50 and percent hemolysis assay_conditions.incubation_time",
                "status": "all_ticketed_rows_set_to_1_h_with_xml_p_86_method_basis",
            },
            {
                "check": "Supplementary Data 10 column E Staphylococcus rows",
                "status": "assigned_ATCC_29213_by_value_provenance_with_source_conflict_preserved",
            },
            {
                "check": "workbook source-cell preservation",
                "status": "all_activity_and_toxicity_rows_checked_against_source_cells",
            },
        ],
        "evidence_paths": [
            "pipeline_v2/deepmine/dbaasp_strict_pilot/papers/PMC12125351/work/activity_evidence/worker2_ticket_xml_section_locator_checks.json",
            "pipeline_v2/deepmine/dbaasp_strict_pilot/papers/PMC12125351/work/activity_evidence/worker2_ticket_source_locator_checks.json",
            reconciliation_path.relative_to(REPO_ROOT).as_posix(),
            validation_path.relative_to(REPO_ROOT).as_posix(),
        ],
        "repaired_artifacts": [path.relative_to(REPO_ROOT).as_posix() for path in ARTIFACT_PATHS],
        "artifacts_written": [
            reconciliation_path.relative_to(REPO_ROOT).as_posix(),
            validation_path.relative_to(REPO_ROOT).as_posix(),
        ],
        "validation_artifacts": [
            validation_path.relative_to(REPO_ROOT).as_posix(),
            reconciliation_path.relative_to(REPO_ROOT).as_posix(),
        ],
        "notes": "Nonterminal owner repair response; worker-6 must re-adjudicate and close if strict gates pass.",
    }
    REWORK_RESPONSES.parent.mkdir(parents=True, exist_ok=True)
    with REWORK_RESPONSES.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(response, ensure_ascii=False) + "\n")


def main() -> None:
    repaired_at = now_iso()
    wb = load_workbook(XLSX_PATH, data_only=True, read_only=True)

    written = []
    repair_counts_by_artifact: dict[str, Any] = {}
    for path in ARTIFACT_PATHS:
        obj = load_json(path)
        repaired_obj, counts = repair_object(obj, repaired_at, wb)
        write_json(path, repaired_obj)
        written.append(path.relative_to(REPO_ROOT).as_posix())
        repair_counts_by_artifact[path.relative_to(REPO_ROOT).as_posix()] = counts

    canonical = load_json(WORK_ACTIVITY)
    validation = validate_obj(canonical, wb)
    validation["paper_id"] = PAPER_ID
    validation["ticket_id"] = TICKET_ID
    validation["validated_at"] = repaired_at
    validation["artifacts_repaired"] = written
    validation["repair_counts_by_artifact"] = repair_counts_by_artifact
    validation["all_ticket_acceptance_checks_passed"] = (
        validation["activity_record_count"] == 130
        and validation["toxicity_record_count"] == 126
        and validation["xlsx_source_cell_rows_checked"] == 256
        and validation["xlsx_source_cell_mismatch_count"] == 0
        and not validation["hemolysis_bad_record_ids"]
        and validation["supp_data10_staph_rows_checked"] == 9
        and not validation["supp_data10_staph_conflict_bad_record_ids"]
        and all(item["has_data3_column_n_value_provenance_match"] for item in validation["sd10_value_provenance_reconciliation"])
        and all(item["data4_column_g_checked_for_reconciliation"] for item in validation["sd10_value_provenance_reconciliation"])
        and validation["invalid_normalization_count"] == 0
        and validation["direct_conversion_conflict_count"] == 0
        and validation["concentration_conflict_count"] == 0
    )

    reconciliation = {
        "paper_id": PAPER_ID,
        "ticket_id": TICKET_ID,
        "validated_at": repaired_at,
        "row_count": validation["supp_data10_staph_rows_checked"],
        "all_data10_values_have_data3_column_n_log10_match": all(
            item["has_data3_column_n_value_provenance_match"] for item in validation["sd10_value_provenance_reconciliation"]
        ),
        "all_data10_values_reconciled_against_data4_column_g": all(
            item["data4_column_g_checked_for_reconciliation"] for item in validation["sd10_value_provenance_reconciliation"]
        ),
        "rows_with_data4_column_g_numeric_overlap": sum(
            1 for item in validation["sd10_value_provenance_reconciliation"] if item["data4_log10_match_count"]
        ),
        "conflict_or_caution_field_required": True,
        "rows": validation["sd10_value_provenance_reconciliation"],
    }

    validation_path = WORK_DIR / "worker2_field_conflict_repair_validation.json"
    reconciliation_path = WORK_DIR / "worker2_sd10_value_provenance_reconciliation.json"
    write_json(validation_path, validation)
    write_json(reconciliation_path, reconciliation)
    response_appended = 0
    if validation["all_ticket_acceptance_checks_passed"]:
        append_rework_response(repaired_at, validation_path, reconciliation_path)
        response_appended = 1

    print(
        "repair_written",
        len(written),
        "validation_pass",
        validation["all_ticket_acceptance_checks_passed"],
        "response_appended",
        response_appended,
    )


if __name__ == "__main__":
    main()
