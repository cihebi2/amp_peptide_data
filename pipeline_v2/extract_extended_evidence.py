#!/usr/bin/env python3
"""pipeline_v2 coverage extension: add PDF tables + full text + supplement tables to the evidence pack.

Root-cause fix addressed: low coverage (74.6% cannot_determine) because values live in PDF text/tables,
supplements, or figures rather than the XML main-text tables. We add reliable extra evidence so v2 can
verify more, and so it can confirm an organism WAS tested (full text) instead of guessing absence.
"""
import json, re, subprocess, sys, warnings
from pathlib import Path

warnings.filterwarnings("ignore")
ROOT = Path(__file__).resolve().parents[1]

VALUE_RE = re.compile(r"^[<>≥≤~=±]?\s*\d")
SKIP_RE = re.compile(r"(?i)^(na|nd|n\.d\.|-|–|—|nt|non-?inhibitory|not\s+(detected|active|inhibitory)|inactive)$")


def is_value_cell(c):
    c = (c or "").strip()
    return bool(c) and (bool(VALUE_RE.match(c)) or bool(SKIP_RE.match(c)))


def grid_to_table(table_index, label, caption, grid):
    grid = [[(c or "").replace("\n", " ").strip() for c in row] for row in grid if any((c or "").strip() for c in row)]
    if not grid:
        return None
    def is_data(row):
        rest = [c for c in row[1:] if c.strip()]
        return len(rest) >= 1 and sum(1 for c in rest if is_value_cell(c)) >= max(1, len(rest) // 2)
    header_rows, dstart = [], len(grid)
    for i, row in enumerate(grid):
        if is_data(row):
            dstart = i
            break
        header_rows.append(row)
    headers = max(header_rows, key=lambda r: sum(1 for c in r if c.strip())) if header_rows else grid[0]
    longform = []
    for ri in range(dstart, len(grid)):
        row = grid[ri]
        if not row:
            continue
        off = 1 if len(row) == len(headers) + 1 else 0
        for ci in range(1, len(row)):
            if not row[ci].strip():
                continue
            hi = ci - off
            longform.append({
                "table_index": table_index, "row_index": ri + 1, "col_index": ci + 1,
                "row_label": row[0], "col_header": headers[hi] if 0 <= hi < len(headers) else f"col{ci}",
                "value": row[ci],
            })
    return {"table_index": table_index, "label": label, "caption": caption,
            "footnotes": [], "header_rows": header_rows, "grid": grid, "longform_cells": longform}


def pdf_tables(pdf_path, start_idx):
    out = []
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            for pi, page in enumerate(pdf.pages):
                for ti, tbl in enumerate(page.extract_tables() or []):
                    t = grid_to_table(start_idx + len(out), f"PDF p{pi+1} table{ti+1}", f"from {pdf_path.name}", tbl)
                    if t and t["longform_cells"]:
                        out.append(t)
    except Exception as e:
        print(f"  pdf_tables err {pdf_path.name}: {e}")
    return out


def pdf_fulltext(pdf_path):
    try:
        r = subprocess.run(["pdftotext", "-q", str(pdf_path), "-"], capture_output=True, text=True, timeout=120)
        return re.sub(r"\n{3,}", "\n\n", r.stdout)[:40000]
    except Exception:
        return ""


def xlsx_tables(path, start_idx):
    out = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            grid = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
            t = grid_to_table(start_idx + len(out), f"SUPP xlsx:{path.name}:{sn}", f"supplement {path.name}", grid)
            if t and t["longform_cells"]:
                out.append(t)
    except Exception as e:
        print(f"  xlsx err {path.name}: {e}")
    return out


def docx_tables(path, start_idx):
    out = []
    try:
        import docx
        d = docx.Document(str(path))
        for di, tb in enumerate(d.tables):
            grid = [[c.text for c in row.cells] for row in tb.rows]
            t = grid_to_table(start_idx + len(out), f"SUPP docx:{path.name}:t{di+1}", f"supplement {path.name}", grid)
            if t and t["longform_cells"]:
                out.append(t)
    except Exception as e:
        print(f"  docx err {path.name}: {e}")
    return out


def find_supplements(paper_id):
    bases = [ROOT / f"papers/{paper_id}/source/supplementary",
             ROOT / f"paper_packets/{paper_id}/extracted/oa_package"]
    files = []
    for b in bases:
        if b.exists():
            for f in b.rglob("*"):
                if f.is_file() and f.suffix.lower() in (".xlsx", ".docx", ".csv", ".pdf"):
                    files.append(f)
    return files


def main():
    paper_id = sys.argv[1]
    pack_path = ROOT / f"pipeline_v2/work/{paper_id}/evidence_pack.json"
    pack = json.loads(pack_path.read_text(encoding="utf-8"))
    nxt = max([t["table_index"] for t in pack["tables"]], default=0) + 1
    added = []
    pdf = ROOT / f"papers/{paper_id}/source/paper.pdf"
    if pdf.exists():
        pt = pdf_tables(pdf, nxt)
        added += pt
        nxt += len(pt)
        pack["fulltext"] = pdf_fulltext(pdf)
    for sup in find_supplements(paper_id)[:8]:
        if sup.suffix.lower() == ".xlsx":
            ad = xlsx_tables(sup, nxt)
        elif sup.suffix.lower() == ".docx":
            ad = docx_tables(sup, nxt)
        elif sup.suffix.lower() == ".pdf":
            ad = pdf_tables(sup, nxt)
        else:
            ad = []
        added += ad
        nxt += len(ad)
    pack["tables"].extend(added)
    pack["extended"] = {"added_tables": len(added), "fulltext_chars": len(pack.get("fulltext", ""))}
    pack_path.write_text(json.dumps(pack, ensure_ascii=False, indent=1), encoding="utf-8")
    cells = sum(len(t["longform_cells"]) for t in added)
    print(f"{paper_id}: +{len(added)} tables (+{cells} cells), fulltext={len(pack.get('fulltext',''))} chars")


if __name__ == "__main__":
    main()
